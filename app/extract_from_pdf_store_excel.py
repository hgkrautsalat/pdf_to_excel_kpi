import PyPDF2
import re
import openpyxl
from openpyxl import Workbook
import os

def extract_text_from_pdf(pdf_path):
    """
    Extrahiert Text aus einer PDF-Datei.

    Args:
        pdf_path (str): Pfad zur PDF-Datei

    Returns:
        str: Extrahierter Text aus der PDF
    """
    text = ""

    try:
        with open(pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)

            # Durchlaufe alle Seiten
            for page in pdf_reader.pages:
                text += page.extract_text()

    except FileNotFoundError:
        print(f"Fehler: Datei '{pdf_path}' nicht gefunden.")
    except Exception as e:
        print(f"Fehler beim Lesen der PDF: {e}")

    return text


def extract_metrics_from_text(text):
    """
    Filtert Kennzahlen aus dem extrahierten Text heraus.

    Args:
        text (str): Der extrahierte PDF-Text

    Returns:
        dict: Dictionary mit gefundenen Kennzahlen
    """

    metrics = {}

    # Suche nach Zahlen mit Währungen (z.B. 1.234,56 €)
    currency_pattern = r'(\d{1,3}(?:\.\d{3})*(?:,\d{2})?)\s*€'
    currencies = re.findall(currency_pattern, text)
    if currencies:
        metrics['Währungsbeträge'] = currencies

    # Suche nach Prozentzahlen (z.B. 25%, 12.5%)
    percent_pattern = r'(\d+(?:[.,]\d+)?)\s*%'
    percentages = re.findall(percent_pattern, text)
    if percentages:
        metrics['Prozentsätze'] = percentages

    # Suche nach Datumswerten (z.B. 01.12.2023, 2023-12-01)
    date_pattern = r'(\d{2}\.\d{2}\.\d{4}|\d{4}-\d{2}-\d{2})'
    dates = re.findall(date_pattern, text)
    if dates:
        metrics['Daten'] = dates

    # Suche nach allgemeinen Zahlen
    number_pattern = r'\b(\d{1,3}(?:\.\d{3})*(?:,\d{2})?)\b'
    numbers = re.findall(number_pattern, text)
    if numbers:
        metrics['Zahlen'] = numbers[:10]  # Limitiere auf erste 10

    return metrics


def write_metrics_to_excel(metrics, excel_path):
    """
    Schreibt die extrahierten Kennzahlen in eine Excel-Datei.

    Args:
        metrics (dict): Dictionary mit Kennzahlen
        excel_path (str): Pfad zur Excel-Datei
    """

    # Erstelle neue Workbook oder lade bestehende
    if os.path.exists(excel_path):
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Kennzahlen"
        # Header erstellen
        sheet['A1'] = "Kategorie"
        sheet['B1'] = "Wert"

    # Bestimme nächste freie Zeile
    row = sheet.max_row + 1

    # Schreibe Kennzahlen in Excel
    for category, values in metrics.items():
        for value in values:
            sheet[f'A{row}'] = category
            sheet[f'B{row}'] = value
            row += 1

    # Speichere Excel-Datei
    try:
        workbook.save(excel_path)
        print(f"Kennzahlen erfolgreich in '{excel_path}' gespeichert.")
    except Exception as e:
        print(f"Fehler beim Speichern der Excel-Datei: {e}")

# Beispielverwendung
if __name__ == "__main__":
    pdf_file = "beispiel.pdf"
    extracted_text = extract_text_from_pdf(pdf_file)
    print(extracted_text)

    metrics = extract_metrics_from_text(extracted_text)
    print(metrics)

    excel_file = "kennzahlen.xlsx"
    write_metrics_to_excel(metrics, excel_file)
    print(f"Kennzahlen wurden in '{excel_file}' gespeichert.")
